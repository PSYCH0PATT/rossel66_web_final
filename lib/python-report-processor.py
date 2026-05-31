import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import shutil
import os
import re
import json
import sys
import time
from collections import defaultdict
from datetime import date, datetime

# Путь к шаблону (lib/templates — в git; корень — legacy для локальной разработки)
def _find_template():
    import unicodedata
    base_dir = os.path.dirname(os.path.abspath(__file__))
    possible_paths = [
        os.path.join(base_dir, 'templates', 'report-mendxza.xlsx'),
        os.path.join(base_dir, '..', 'Отчёт MENDXZA.xlsx'),
        os.path.join(base_dir, '..', 'Отчёт MENDXZA.xlsx'),
        'Отчёт MENDXZA.xlsx',
        'Отчёт MENDXZA.xlsx'
    ]
    for p in possible_paths:
        if os.path.exists(p):
            return p
        try:
            norm_p = unicodedata.normalize('NFC', p)
            if os.path.exists(norm_p):
                return norm_p
            norm_p = unicodedata.normalize('NFD', p)
            if os.path.exists(norm_p):
                return norm_p
        except Exception:
            pass
            
    # Поиск по шаблону MENDXZA в родительской директории
    try:
        parent_dir = os.path.abspath(os.path.join(base_dir, '..'))
        if os.path.exists(parent_dir):
            for f in os.listdir(parent_dir):
                if 'MENDXZA' in f and f.endswith('.xlsx'):
                    return os.path.join(parent_dir, f)
    except Exception:
        pass
                
    return os.path.join(base_dir, 'templates', 'report-mendxza.xlsx')

TEMPLATE_PATH = os.environ.get('TEMPLATE_PATH') or _find_template()

def _sanitize_filename(name):
    """Безопасное имя файла отчёта (Linux/Timeweb)."""
    safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', str(name))
    safe = safe.strip().strip('.')
    return safe or 'artist'

# Ожидаемые колонки отчёта и возможные варианты названий в Excel
COLUMN_ALIASES = {
    'Код': ['Код', 'код', 'Код трека', 'ISRC', 'isrc'],
    'Исполнитель': ['Исполнитель', 'исполнитель', 'Artist'],
    'Наименование': ['Наименование', 'наименование', 'Название', 'Трек'],
    'Альбом': ['Альбом', 'альбом', 'Release'],
    'Количество': ['Количество', 'количество', 'Кол-во', 'Прослушивания'],
    'Сумма, руб.': ['Сумма, руб.', 'Сумма,руб.', 'Сумма (руб.)', 'Сумма руб.', 'Сумма руб', 'Сумма, руб', 'Сумма', 'сумма, руб.', 'Сумма (руб)'],
}

CANONICAL_FROM_MAPPING = {
    'isrc_column': 'Код',
    'artist_column': 'Исполнитель',
    'track_name_column': 'Наименование',
    'album_name_column': 'Альбом',
    'plays_column': 'Количество',
    'amount_column': 'Сумма, руб.',
}

REQUIRED_STATEMENT_COLUMNS = list(CANONICAL_FROM_MAPPING.values())

QUARTER_LABELS = {
    'Q1': (1, '01.01.{year}', '31.03.{year}'),
    'Q2': (2, '01.04.{year}', '30.06.{year}'),
    'Q3': (3, '01.07.{year}', '30.09.{year}'),
    'Q4': (4, '01.10.{year}', '31.12.{year}'),
}

def _format_quarter_labels(quarter, year):
    """Returns (B8 text, D8 period text) for the report summary sheet."""
    key = str(quarter).strip().upper()
    if key not in QUARTER_LABELS:
        raise ValueError(f"Неизвестный квартал: {quarter}. Ожидается Q1, Q2, Q3 или Q4.")
    num, start_tpl, end_tpl = QUARTER_LABELS[key]
    start = start_tpl.format(year=year)
    end = end_tpl.format(year=year)
    return f"{num} квартал {year} г.", f"Период с {start} по {end}"

def _parse_approval_date(raw_value=None):
    """Parse REPORT_APPROVAL_DATE env (YYYY-MM-DD) or default to today."""
    value = raw_value if raw_value is not None else os.environ.get('REPORT_APPROVAL_DATE')
    if value:
        value = str(value).strip()
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            pass
    return date.today()

def _set_cell(ws, cell_ref, value, number_format=None):
    """Записывает значение в ячейку; если ячейка объединена — в левую верхнюю ячейку диапазона."""
    target_cell = None
    try:
        cell = ws[cell_ref]
        if type(cell).__name__ == 'MergedCell':
            col_letter, row = coordinate_from_string(cell_ref)
            col = column_index_from_string(col_letter)
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                    target_cell = ws.cell(row=mr.min_row, column=mr.min_col)
                    break
        else:
            target_cell = cell
    except AttributeError:
        col_letter, row = coordinate_from_string(cell_ref)
        col = column_index_from_string(col_letter)
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                target_cell = ws.cell(row=mr.min_row, column=mr.min_col)
                break
        if not target_cell:
            raise

    if target_cell:
        target_cell.value = value
        if number_format:
            target_cell.number_format = number_format

def _normalize_statement_columns(df):
    """Приводит названия колонок к ожидаемым (разные форматы выгрузки)."""
    rename = {}
    for col in df.columns:
        c = str(col).strip() if pd.notna(col) else ''
        for canonical, aliases in COLUMN_ALIASES.items():
            if c in aliases or c == canonical:
                rename[col] = canonical
                break
    return df.rename(columns=rename)

def _col_letter_to_index(letter):
    """Excel column letter to 0-based index (A=0, B=1, ..., Z=25)."""
    if not letter or not isinstance(letter, str):
        return None
    letter = letter.strip().upper()
    if not letter.isalpha():
        return None
    idx = 0
    for char in letter:
        idx = idx * 26 + (ord(char) - ord('A') + 1)
    return idx - 1

def _cell_str(val):
    if pd.isna(val):
        return ''
    return str(val).strip()

def _row_header_score(row, column_mapping=None):
    """Higher score = more likely a header row."""
    aliases_flat = set()
    for aliases in COLUMN_ALIASES.values():
        for alias in aliases:
            aliases_flat.add(alias.lower())

    score = 0
    for val in row:
        cell = _cell_str(val).lower()
        if not cell:
            continue
        if cell in aliases_flat:
            score += 2
        elif any(alias in cell or cell in alias for alias in aliases_flat if len(alias) > 3):
            score += 1

    if column_mapping:
        for field in CANONICAL_FROM_MAPPING:
            letter = column_mapping.get(field)
            if not letter:
                continue
            idx = _col_letter_to_index(letter)
            if idx is not None and idx < len(row):
                cell = _cell_str(row.iloc[idx]).lower()
                if cell and not cell.replace('.', '').replace(',', '').replace('-', '').isdigit():
                    score += 1
    return score

def _find_header_row(raw_df, column_mapping=None, min_score=3):
    best_row = None
    best_score = 0
    for i in range(min(40, len(raw_df))):
        score = _row_header_score(raw_df.iloc[i], column_mapping)
        if score > best_score:
            best_score = score
            best_row = i
    if best_score >= min_score:
        return best_row
    return None

def _load_column_mapping(path):
    if not path or not os.path.isfile(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return {key: value for key, value in data.items() if value}

def _coerce_numeric_columns(df):
    for col in ['Количество', 'Сумма, руб.']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df

def _load_statement_df(statement_path, column_mapping=None):
    """Load distributor statement; supports title rows and UI column letter mapping."""
    xl = pd.ExcelFile(statement_path)
    sheet_name = 'TDSheet' if 'TDSheet' in xl.sheet_names else xl.sheet_names[0]

    if column_mapping:
        raw = pd.read_excel(statement_path, sheet_name=sheet_name, header=None)
        header_row = _find_header_row(raw, column_mapping)
        if header_row is None:
            header_row = 0

        data = raw.iloc[header_row + 1:].copy().reset_index(drop=True)
        result = pd.DataFrame()
        for field, canonical in CANONICAL_FROM_MAPPING.items():
            letter = column_mapping.get(field)
            if not letter:
                continue
            idx = _col_letter_to_index(letter)
            if idx is None or idx >= raw.shape[1]:
                raise ValueError(
                    f"Столбец {letter} ({canonical}) вне диапазона файла "
                    f"(всего столбцов: {raw.shape[1]})"
                )
            result[canonical] = data.iloc[:, idx]

        result = result.dropna(how='all')
        result = result[
            ~result.apply(
                lambda row: all(_cell_str(v) == '' for v in row),
                axis=1,
            )
        ]
        return _coerce_numeric_columns(result)

    statement_df = pd.read_excel(statement_path, sheet_name=sheet_name)
    statement_df = _normalize_statement_columns(statement_df)
    missing = [col for col in REQUIRED_STATEMENT_COLUMNS if col not in statement_df.columns]
    if missing:
        raw = pd.read_excel(statement_path, sheet_name=sheet_name, header=None)
        header_row = _find_header_row(raw)
        if header_row is not None:
            statement_df = pd.read_excel(statement_path, sheet_name=sheet_name, header=header_row)
            statement_df = _normalize_statement_columns(statement_df)
            missing = [col for col in REQUIRED_STATEMENT_COLUMNS if col not in statement_df.columns]

    if missing:
        raise ValueError(
            f"В файле отчёта не найдены колонки: {missing}. "
            f"Есть колонки: {list(statement_df.columns)}"
        )
    return _coerce_numeric_columns(statement_df)

def _is_empty_report_value(val):
    if val is None:
        return True
    s = str(val).strip()
    return s == '' or s == '-'

REPORT_FIELD_LABELS = {
    'fio': 'ФИО',
    'contract': 'Номер договора',
    'percentage': 'Процент',
}

def _missing_report_fields(user):
    missing = []
    if _is_empty_report_value(user.get('fio')):
        missing.append('fio')
    if _is_empty_report_value(user.get('contract')):
        missing.append('contract')
    if user.get('percentage') is None or user.get('percentage') == '':
        missing.append('percentage')
    return missing

def _emit_incomplete_report_json(skipped_incomplete):
    if skipped_incomplete:
        payload = json.dumps({"incompleteArtists": skipped_incomplete}, ensure_ascii=False)
        print("REPORT_INCOMPLETE_JSON:" + payload)

def _collect_statement_artists(statement_df, match_list):
    """Artists from Supabase that appear in the uploaded statement file."""
    artists = set()
    for _, row in statement_df.iterrows():
        for artist in extract_artists_from_track(row['Исполнитель'], match_list):
            artists.add(artist)
    return artists

def _filter_incomplete_for_statement(skipped_incomplete, statement_artists):
    return [item for item in skipped_incomplete if item['name'] in statement_artists]

def get_artists_list_from_users(users_file):
    """Читает данные артистов из Prisma export snapshot (/tmp/users_export_*.json)."""
    skipped_incomplete = []
    try:
        if not os.path.exists(users_file):
            print(f"Файл {users_file} не найден")
            return {}, [], skipped_incomplete
        
        with open(users_file, 'r', encoding='utf-8') as f:
            users_data = json.load(f)
        
        artists_dict = {}
        match_list = []  # (canonical_key, [name, username, ...]) для сопоставления по строке исполнителя
        for user in users_data:
            if user.get('role') != 'artist':
                continue
            canonical = user.get('name') or user.get('username')
            if not canonical:
                continue
            aliases = [a for a in (user.get('name'), user.get('username')) if a]
            missing = _missing_report_fields(user)
            if missing:
                labels = [REPORT_FIELD_LABELS.get(f, f) for f in missing]
                print(f"⚠️  Пропущен артист {canonical}: не хватает данных для отчёта ({', '.join(labels)})")
                skipped_incomplete.append({"name": canonical, "missingFields": missing})
                match_list.append((canonical, aliases))
                continue
            
            percentage = user.get('percentage')
            artists_dict[canonical] = [
                user.get('fio') or user.get('name') or '',
                user.get('fioShort') or user.get('fio') or user.get('name') or '',
                user.get('contract') or '',
                str(percentage),
                user.get('id')
            ]
            match_list.append((canonical, aliases))
        
        print(f"✅ Загружено {len(artists_dict)} артистов из {users_file}")
        return artists_dict, match_list, skipped_incomplete
    except Exception as e:
        print(f"Ошибка при чтении {users_file}: {e}")
        return {}, [], skipped_incomplete

def get_royalty_shares():
    # Возвращаем пустой словарь (как в оригинале)
    return {}

def get_royalty_shares_from_file(file_path):
    """Читает доли роялти из переданного файла"""
    try:
        df = pd.read_excel(file_path)
        shares = {}
        for _, row in df.iloc[1:].iterrows():
            if pd.notna(row[0]):
                track_code = str(row[0])
                performer = str(row[3]) if pd.notna(row[3]) else None
                percent = row[4] if pd.notna(row[4]) else None
                if track_code and performer and percent:
                    if track_code not in shares:
                        shares[track_code] = {}
                    try:
                        percent_str = str(percent)
                        percent_value = float(percent_str.replace('%', '')) / 100
                        shares[track_code][performer] = percent_value
                    except (ValueError, TypeError):
                        continue
        return shares
    except Exception as e:
        print(f"Ошибка при чтении файла долей: {e}")
        return {}

def extract_artists_from_track(artist_str, match_list):
    """Извлекает список артистов из строки исполнителя. Возвращает только канонические имена (без дубликатов по регистру)."""
    if not isinstance(artist_str, str):
        artist_str = str(artist_str)
    found = []
    for canonical, aliases in match_list:
        for alias in aliases:
            if alias and re.search(re.escape(alias), artist_str, re.IGNORECASE):
                found.append(canonical)
                break
    return found

def get_royalty_shares_from_tracks(releases_file):
    """Загружает доли роялти из треков в указанном releases.json"""
    try:
        if not os.path.exists(releases_file):
            return {}
        
        with open(releases_file, 'r', encoding='utf-8') as f:
            releases_data = json.load(f)
        
        track_shares = {}
        for release in releases_data:
            if not release.get('tracks'):
                continue
            
            for track in release.get('tracks', []):
                isrc = track.get('isrc')
                royalty_shares = track.get('royaltyShares')
                
                if isrc and royalty_shares and isinstance(royalty_shares, dict):
                    # Конвертируем проценты в доли (0-1)
                    if isrc not in track_shares:
                        track_shares[isrc] = {}
                    for artist_name, percentage in royalty_shares.items():
                        if percentage is not None and percentage > 0:
                            try:
                                track_shares[isrc][artist_name] = float(percentage) / 100.0
                            except (ValueError, TypeError):
                                continue
        
        if track_shares:
            print(f"✅ Загружено {len(track_shares)} треков с долями роялти из {releases_file}")
        return track_shares
    except Exception as e:
        print(f"Ошибка при чтении долей из треков: {e}")
        return {}

def calculate_artist_share(track_code, artist, all_artists_in_track, artists_data, royalty_shares, track_royalty_shares=None):
    """Рассчитывает долю артиста в треке"""
    # 1. Если артист один в треке - получает 100%
    if len(all_artists_in_track) == 1:
        return 1.0
    
    # 2. ВЫСШИЙ ПРИОРИТЕТ: Доли из треков (releases.json)
    if track_royalty_shares and track_code in track_royalty_shares:
        track_shares = track_royalty_shares[track_code]
        if artist in track_shares:
            return track_shares[artist]
    
    # 3. ПРИОРИТЕТ: Проверяем файл "Доли роялти" (точные доли для конкретного трека)
    if track_code in royalty_shares:
        track_shares = royalty_shares[track_code]
        if artist in track_shares:
            return track_shares[artist]
    
    # 4. Если все артисты в треке зарегистрированы - делим поровну
    if all(a in artists_data for a in all_artists_in_track):
        return 1.0 / len(all_artists_in_track)
    
    # 5. Считаем только наших артистов (из artists_data)
    our_artists_count = sum(1 for a in all_artists_in_track if a in artists_data)
    if our_artists_count > 0:
        return 1.0 / our_artists_count
    
    # 6. Если артист не найден - получает 0%
    return 0.0

def process_file(statement_path, quarter, year, users_file, releases_file, reports_dir, metadata_json_path, royalty_file_path=None):
    """Обработка файла с единым хранилищем отчетов"""
    if not os.path.isfile(TEMPLATE_PATH):
        raise FileNotFoundError(
            f"Шаблон отчёта не найден: {TEMPLATE_PATH}. "
            f"Ожидается lib/templates/report-mendxza.xlsx в репозитории."
        )

    # Создаем единую папку для всех отчетов в tmp
    os.makedirs(reports_dir, exist_ok=True)
    
    # Загружаем список пользователей для проверки регистрации
    registered_users = set()
    if os.path.exists(users_file):
        with open(users_file, 'r', encoding='utf-8') as f:
            users_data = json.load(f)
            registered_users = {user.get('username') for user in users_data if user.get('username')}
            for user in users_data:
                if user.get('name'):
                    registered_users.add(user.get('name'))
    
    column_mapping_path = os.environ.get('COLUMN_MAPPING_PATH')
    column_mapping = _load_column_mapping(column_mapping_path)
    if column_mapping:
        print(f"📋 Маппинг столбцов: {column_mapping}")

    statement_df = _load_statement_df(statement_path, column_mapping)

    missing = [col for col in REQUIRED_STATEMENT_COLUMNS if col not in statement_df.columns]
    if missing:
        raise ValueError(
            f"В файле отчёта не найдены колонки: {missing}. "
            f"Есть колонки: {list(statement_df.columns)}"
        )

    # Загружаем данные артистов
    artists_data, match_list, skipped_incomplete = get_artists_list_from_users(users_file)
    statement_artists = _collect_statement_artists(statement_df, match_list)
    incomplete_in_statement = _filter_incomplete_for_statement(skipped_incomplete, statement_artists)
    
    if not artists_data:
        print(f"⚠️  Не найдено артистов с полными данными для отчёта в {users_file}")
        _emit_incomplete_report_json(incomplete_in_statement)
        return []
        
    quarter_label, period_label = _format_quarter_labels(quarter, year)
    approval_date = _parse_approval_date()
        
    # Загружаем доли роялти из треков (высший приоритет)
    track_royalty_shares = get_royalty_shares_from_tracks(releases_file)
    
    # Загружаем доли роялти из файла (если передан) - используется как fallback
    if royalty_file_path and os.path.exists(royalty_file_path):
        royalty_shares = get_royalty_shares_from_file(royalty_file_path)
    else:
        royalty_shares = get_royalty_shares()
    
    artists_tracks = defaultdict(lambda: defaultdict(lambda: {'Количество': 0, 'Сумма, руб.': 0, 'Доля': 0}))
    for _, row in statement_df.iterrows():
        track_code = row['Код']
        artist_str = row['Исполнитель']
        track_artists = extract_artists_from_track(artist_str, match_list)
        if not track_artists:
            continue
        for artist in track_artists:
            share = calculate_artist_share(track_code, artist, track_artists, artists_data, royalty_shares, track_royalty_shares)
            amount_share = row['Сумма, руб.'] * share
            track_key = (track_code, row['Исполнитель'], row['Наименование'], row['Альбом'])
            artists_tracks[artist][track_key]['Количество'] += row['Количество']
            artists_tracks[artist][track_key]['Сумма, руб.'] += amount_share
            artists_tracks[artist][track_key]['Доля'] = share * 100

    created_files = []
    reports_metadata = []
    
    for artist, tracks in artists_tracks.items():
        if artist not in artists_data:
            print(f"⚠️  Пропущен артист {artist}: не найден в списке артистов с процентом")
            continue
        
        safe_name = _sanitize_filename(artist)
        artist_file_path = os.path.join(reports_dir, f'{safe_name}.xlsx')
        is_registered = artist in registered_users
        
        print(f"Создаем отчет для {'зарегистрированного' if is_registered else 'незарегистрированного'} артиста: {artist}")
        
        shutil.copy(TEMPLATE_PATH, artist_file_path)
        wb = load_workbook(artist_file_path)
        summary_sheet_name = 'Итог' if 'Итог' in wb.sheetnames else ('Краткая сводка' if 'Краткая сводка' in wb.sheetnames else wb.sheetnames[0])
        ws = wb[summary_sheet_name]
        total_amount_raw = sum(track['Сумма, руб.'] for track in tracks.values())
        total_amount = round(total_amount_raw, 2)
        
        percentage_str = str(artists_data[artist][3]).strip().replace('%', '').replace(',', '.')
        try:
            artist_percent = float(percentage_str) / 100.0
            percentage_text = f"{percentage_str.rstrip('0').rstrip('.') if '.' in percentage_str else percentage_str}%"
        except ValueError:
            artist_percent = 1.0
            percentage_text = "100%"
            
        final_amount = round(total_amount * artist_percent, 2)
        
        _set_cell(ws, 'B8', quarter_label)                     # B8 Квартал
        _set_cell(ws, 'D8', period_label)                      # D8 Период
        _set_cell(ws, 'B10', artist)                           # B10 Артист
        _set_cell(ws, 'B4', artists_data[artist][2])           # B4 Договор
        _set_cell(ws, 'B6', artists_data[artist][0])           # B6 ФИО полное
        _set_cell(ws, 'E14', total_amount, '0.00')             # E14 Доход
        _set_cell(ws, 'D15', percentage_text)                  # D15 Процент
        _set_cell(ws, 'E15', final_amount, '0.00')             # E15 Доход после вычета
        _set_cell(ws, 'F15', 'руб.')                           # F15 Валюта
        _set_cell(ws, 'E26', final_amount, '0.00')             # E26 Начислено вознаграждение
        _set_cell(ws, 'E28', final_amount, '0.00')             # E28 К выплате
        _set_cell(ws, 'C30', approval_date)                    # C30 Дата утверждения
        _set_cell(ws, 'D32', artists_data[artist][0])          # D32 Лицензиар ФИО
        _set_cell(ws, 'E37', artists_data[artist][1])          # E37 ФИО кратко
        ws_artist = wb.create_sheet(title=artist)
        ws_artist.append(['Код', 'Исполнитель', 'Наименование', 'Альбом', 'Количество', 'Сумма, руб.', 'Доля, %'])
        total_quantity = 0
        total_amount_sheet = 0
        for (track_code, performer, name, album), track in tracks.items():
            track_sum = round(track['Сумма, руб.'], 2)
            track_share = round(track['Доля'], 2)
            data = [track_code, performer, name, album, track['Количество'], track_sum, track_share]
            ws_artist.append(data)
            total_quantity += track['Количество']
            total_amount_sheet += track_sum
        ws_artist.append(['Итого', '', '', '', total_quantity, round(total_amount_sheet, 2), ''])
        wb.save(artist_file_path)
        created_files.append(artist_file_path)
        
        total_plays = sum(track['Количество'] for track in tracks.values())
        
        report_id = f"report_{artist}_{quarter}_{year}_{int(time.time())}_{int(hash(artist) % 1000)}"
        
        report_metadata = {
            "id": report_id,
            "artistId": artists_data[artist][4] if (is_registered and len(artists_data[artist]) > 4) else None,
            "artistName": artist,
            "quarter": quarter,
            "year": year,
            "fileName": f"{safe_name}.xlsx",
            "filePath": artist_file_path,
            "uploadDate": time.strftime('%Y-%m-%dT%H:%M:%S.000Z'),
            "status": "processed",
            "totalPlays": total_plays,
            "totalAmount": final_amount,
            "isRegistered": is_registered,
            "isSigned": False,
            "isPaid": False
        }
        
        reports_metadata.append(report_metadata)
    
    with open(metadata_json_path, 'w', encoding='utf-8') as f:
        json.dump(reports_metadata, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Сохранено {len(reports_metadata)} отчетов в метаданные {metadata_json_path}")
    print(f"Всего создано файлов: {len(created_files)}")
    print(f"Зарегистрированных артистов: {sum(1 for artist in artists_tracks.keys() if artist in registered_users)}")
    print(f"Незарегистрированных артистов: {sum(1 for artist in artists_tracks.keys() if artist not in registered_users)}")
    if incomplete_in_statement:
        _emit_incomplete_report_json(incomplete_in_statement)
    return created_files


if __name__ == "__main__":
    if len(sys.argv) < 8:
        print("Использование: python python-report-processor.py <путь_к_файлу> <квартал> <год> <путь_к_users.json> <путь_к_releases.json> <reports_out_dir> <metadata_json_path> [путь_к_файлу_долей]")
        sys.exit(1)
    
    file_path = sys.argv[1]
    quarter = sys.argv[2]
    year = int(sys.argv[3])
    users_file = sys.argv[4]
    releases_file = sys.argv[5]
    reports_dir = sys.argv[6]
    metadata_json_path = sys.argv[7]
    royalty_file_path = sys.argv[8] if len(sys.argv) > 8 else None
    
    try:
        created_files = process_file(
            file_path, quarter, year, 
            users_file, releases_file, 
            reports_dir, metadata_json_path, 
            royalty_file_path
        )
        if not created_files:
            print("Обработка завершена без созданных отчётов")
            sys.exit(1)
        print("Обработка завершена успешно!")
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
